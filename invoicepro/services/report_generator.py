"""
Report Generator Service
========================
Aggregates invoice, payment, GST, customer, and product data into
report structures that can be rendered in the UI or exported.
"""

import csv
import io
from datetime import date, timedelta
from decimal import Decimal

from sqlalchemy import func

from invoicepro.database.database import db
from invoicepro.database.models import Customer, Invoice, InvoiceItem, Payment, Product


# ── Date helpers ─────────────────────────────────────────────────────────────

def _date_range(period: str, date_from=None, date_to=None):
    today = date.today()
    if period == "today":
        return today, today
    if period == "week":
        start = today - timedelta(days=today.weekday())
        return start, today
    if period == "month":
        return today.replace(day=1), today
    if period == "year":
        return today.replace(month=1, day=1), today
    if period == "custom" and date_from and date_to:
        return date_from, date_to
    # default: this month
    return today.replace(day=1), today


# ── Core queries ─────────────────────────────────────────────────────────────

def revenue_summary(user_id: int, period="month", date_from=None, date_to=None) -> dict:
    start, end = _date_range(period, date_from, date_to)
    q = Invoice.query.filter(
        Invoice.user_id == user_id,
        Invoice.invoice_date >= start,
        Invoice.invoice_date <= end,
    )
    rows = q.all()

    total = sum(Decimal(r.total_amount or 0) for r in rows)
    paid = sum(Decimal(r.amount_paid or 0) for r in rows)
    pending = sum(
        Decimal(r.balance_due) for r in rows
        if r.status in {"Pending", "Sent", "Draft"}
    )
    overdue = sum(
        Decimal(r.balance_due) for r in rows
        if r.status == "Overdue"
    )

    return {
        "period": period,
        "date_from": start,
        "date_to": end,
        "total_revenue": total,
        "total_paid": paid,
        "total_pending": pending,
        "total_overdue": overdue,
        "invoice_count": len(rows),
        "invoices": rows,
    }


def monthly_revenue_chart(user_id: int, year: int | None = None) -> dict:
    year = year or date.today().year
    results = (
        db.session.query(
            func.strftime("%m", Invoice.invoice_date).label("month"),
            func.sum(Invoice.total_amount).label("revenue"),
            func.sum(Invoice.amount_paid).label("paid"),
        )
        .filter(
            Invoice.user_id == user_id,
            func.strftime("%Y", Invoice.invoice_date) == str(year),
        )
        .group_by("month")
        .all()
    )
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    revenue = [0.0] * 12
    paid = [0.0] * 12
    for row in results:
        idx = int(row.month) - 1
        revenue[idx] = float(row.revenue or 0)
        paid[idx] = float(row.paid or 0)
    return {"labels": months, "revenue": revenue, "paid": paid}


def top_customers(user_id: int, limit=5) -> list[dict]:
    rows = (
        db.session.query(Customer, func.sum(Invoice.total_amount).label("total"))
        .join(Invoice, Invoice.customer_id == Customer.id)
        .filter(Customer.user_id == user_id)
        .group_by(Customer.id)
        .order_by(func.sum(Invoice.total_amount).desc())
        .limit(limit)
        .all()
    )
    return [
        {"name": c.name, "total": float(t or 0), "invoice_count": len(c.invoices)}
        for c, t in rows
    ]


def best_selling_products(user_id: int, limit=5) -> list[dict]:
    rows = (
        db.session.query(
            Product,
            func.sum(InvoiceItem.quantity).label("qty"),
            func.sum(InvoiceItem.line_total).label("revenue"),
        )
        .join(InvoiceItem, InvoiceItem.product_id == Product.id)
        .filter(Product.user_id == user_id)
        .group_by(Product.id)
        .order_by(func.sum(InvoiceItem.line_total).desc())
        .limit(limit)
        .all()
    )
    return [
        {"name": p.name, "quantity": float(q or 0), "revenue": float(r or 0)}
        for p, q, r in rows
    ]


def gst_summary(user_id: int, period="month", date_from=None, date_to=None) -> dict:
    start, end = _date_range(period, date_from, date_to)
    invoices = Invoice.query.filter(
        Invoice.user_id == user_id,
        Invoice.invoice_date >= start,
        Invoice.invoice_date <= end,
        Invoice.status != "Cancelled",
    ).all()

    cgst = sum(Decimal(i.cgst_amount or 0) for i in invoices)
    sgst = sum(Decimal(i.sgst_amount or 0) for i in invoices)
    igst = sum(Decimal(i.igst_amount or 0) for i in invoices)
    taxable = sum(Decimal(i.taxable_amount or 0) for i in invoices)

    return {
        "date_from": start,
        "date_to": end,
        "taxable_amount": taxable,
        "cgst": cgst,
        "sgst": sgst,
        "igst": igst,
        "total_tax": cgst + sgst + igst,
        "invoice_count": len(invoices),
    }


def customer_report(user_id: int) -> list[dict]:
    customers = Customer.query.filter_by(user_id=user_id).order_by(Customer.name).all()
    return [
        {
            "name": c.name,
            "email": c.email or "",
            "invoices": len(c.invoices),
            "total_billed": float(c.total_billed),
            "total_paid": float(c.total_paid),
            "outstanding": float(c.total_outstanding),
        }
        for c in customers
    ]


def product_report(user_id: int) -> list[dict]:
    return best_selling_products(user_id, limit=100)


# ── CSV Export ────────────────────────────────────────────────────────────────

def export_invoices_csv(user_id: int, period="month", date_from=None, date_to=None) -> str:
    summary = revenue_summary(user_id, period, date_from, date_to)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Invoice #", "Customer", "Date", "Due Date", "Status",
                     "Currency", "Subtotal", "Discount", "Taxable",
                     "CGST", "SGST", "IGST", "Total", "Paid", "Balance"])
    for inv in summary["invoices"]:
        writer.writerow([
            inv.invoice_number, inv.customer.name,
            inv.invoice_date, inv.due_date, inv.status, inv.currency,
            inv.subtotal, inv.discount_total, inv.taxable_amount,
            inv.cgst_amount, inv.sgst_amount, inv.igst_amount,
            inv.total_amount, inv.amount_paid, inv.balance_due,
        ])
    return output.getvalue()


def export_invoices_excel(user_id: int, period="month", date_from=None, date_to=None) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    summary = revenue_summary(user_id, period, date_from, date_to)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoices"

    headers = ["Invoice #", "Customer", "Date", "Due Date", "Status",
               "Currency", "Subtotal", "Discount", "Taxable",
               "CGST", "SGST", "IGST", "Total", "Paid", "Balance"]
    header_fill = PatternFill(start_color="1A233B", end_color="1A233B", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    for row_idx, inv in enumerate(summary["invoices"], 2):
        ws.append([
            inv.invoice_number, inv.customer.name,
            str(inv.invoice_date), str(inv.due_date), inv.status, inv.currency,
            float(inv.subtotal), float(inv.discount_total), float(inv.taxable_amount),
            float(inv.cgst_amount), float(inv.sgst_amount), float(inv.igst_amount),
            float(inv.total_amount), float(inv.amount_paid), float(inv.balance_due),
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
